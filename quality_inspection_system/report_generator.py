import os
import csv
import json
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

def generate_pdf_report(inspection_data):
    insp_id = inspection_data["id"]
    filename = f"report_{insp_id}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=12
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569')
    )
    
    elements = []
    
    # Title & Metadata
    elements.append(Paragraph("AUTOMATED QUALITY INSPECTION REPORT", title_style))
    elements.append(Paragraph(f"Inspection ID: {insp_id} | Date: {inspection_data['timestamp']}", header_style))
    elements.append(Spacer(1, 15))
    
    # Summary Table
    result_color = colors.HexColor('#16A34A') if inspection_data['result'] == 'PASS' else colors.HexColor('#DC2626')
    
    table_data = [
        ["Parameter", "Inspection Value"],
        ["Product Type", str(inspection_data.get('product_type', 'N/A'))],
        ["Inspection Status", inspection_data['result']],
        ["Defect Category", str(inspection_data.get('defect_type', inspection_data.get('defect', 'None')))],
        ["Confidence Score", f"{float(inspection_data['confidence']) * 100:.1f}%"],
        ["Processing Latency", f"{inspection_data['total_time_ms']} ms"],
        ["Width (mm)", f"{inspection_data.get('width_mm', 0):.2f} mm"],
        ["Height (mm)", f"{inspection_data.get('height_mm', 0):.2f} mm"],
        ["Area (mm²)", f"{inspection_data.get('area_mm2', 0):.2f} mm²"]
    ]
    
    t = Table(table_data, colWidths=[200, 300])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#0284C7')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
        ('TEXTCOLOR', (1, 2), (1, 2), result_color)
    ]))
    
    elements.append(t)
    elements.append(Spacer(1, 20))
    
    # 15-Stage Diagnostics
    elements.append(Paragraph("15-Stage OpenCV Computer Vision Pipeline Diagnostic Log", styles['Heading2']))
    elements.append(Spacer(1, 8))
    
    stages_data = [["#", "Stage Name", "Latency", "Status", "Technical Diagnostic"]]
    for stage in inspection_data.get('stages', []):
        stages_data.append([
            str(stage.get('number', '')),
            str(stage.get('name', '')),
            f"{stage.get('time_ms', 0)} ms",
            str(stage.get('status', 'PASS')),
            str(stage.get('explanation', ''))
        ])
        
    st_table = Table(stages_data, colWidths=[25, 110, 55, 50, 260])
    st_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0'))
    ]))
    
    elements.append(st_table)
    
    doc.build(elements)
    return filepath

def generate_csv_report(inspection_data):
    insp_id = inspection_data["id"]
    filename = f"report_{insp_id}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    with open(filepath, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Inspection ID", "Timestamp", "Filename", "Product Type", "Result", "Defect", "Confidence", "Width_mm", "Height_mm", "Area_mm2", "Latency_ms"])
        writer.writerow([
            inspection_data["id"],
            inspection_data["timestamp"],
            inspection_data["filename"],
            inspection_data.get("product_type", "Unknown"),
            inspection_data["result"],
            inspection_data.get("defect_type", inspection_data.get("defect", "None")),
            inspection_data["confidence"],
            inspection_data.get("width_mm", 0),
            inspection_data.get("height_mm", 0),
            inspection_data.get("area_mm2", 0),
            inspection_data["total_time_ms"]
        ])
    return filepath

def generate_excel_report(inspection_data):
    insp_id = inspection_data["id"]
    filename = f"report_{insp_id}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspection Summary"
    
    headers = ["Inspection ID", "Timestamp", "Product Type", "Result", "Defect", "Confidence", "Width (mm)", "Height (mm)", "Latency (ms)"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    row_data = [
        inspection_data["id"],
        inspection_data["timestamp"],
        inspection_data.get("product_type", "Unknown"),
        inspection_data["result"],
        inspection_data.get("defect_type", inspection_data.get("defect", "None")),
        f"{float(inspection_data['confidence']) * 100:.1f}%",
        inspection_data.get("width_mm", 0),
        inspection_data.get("height_mm", 0),
        inspection_data["total_time_ms"]
    ]
    ws.append(row_data)
    
    wb.save(filepath)
    return filepath

def generate_json_report(inspection_data):
    insp_id = inspection_data["id"]
    filename = f"report_{insp_id}.json"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(inspection_data, f, indent=2)
        
    return filepath
